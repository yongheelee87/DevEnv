import datetime
from enum import Enum, auto
import threading
import openpyxl

from Lib.Inst.visaLib import *
from Lib.Inst.telnetLib import *
from Lib.Inst import *

# Command
CMD_LIST = ['CMD', 'Time[sec]', 'Env_Val', 'Servo_Speed', 'Alpha_AngleCh1', 'outMotReqCh1', 'stMotOffReqCh1', 'stMotCtrlTarCh1',
            'nMotCtrlAbsLimCh1', 'CurrTrig1stCh1', 'CurrTrig2ndCh1', 'Alpha_AngleCh2', 'outMotReqCh2', 'stMotOffReqCh2',
            'stMotCtrlTarCh2', 'nMotCtrlAbsLimCh2', 'CurrTrig1stCh2', 'CurrTrig2ndCh2', 'Test_Flag']
MEASURE_LIST = ['Date Time', 'Time[sec]', 'Servo_Speed', 'outMotReqCh1', 'outMotReqCh2', 'Alpha_AngleCh1', '',
                'kt', 'dc_current', 'dc_voltage', 'dynamo_torque', 'friction_torque', 'dynamo_speed', 'dynamo_volt',
                'dynamo_curr', 'current_p1', 'current_p2', 'current_p3', 'voltage_p1', 'voltage_p2', 'voltage_p3',
                'power_ratio_p1', 'power_ratio_p2', 'power_ratio_p3', 'temp_ch1', 'current_value_ch1', 'battery_voltage_gate_driver',
                'motor_operating_status_ch1', 'motor_speed_value_ch1']


class CmdState(Enum):
    # Running State #
    READY = auto()
    RUNNING = auto()
    NEXT = auto()
    DONE = auto()
    TEMP_WAIT_MODE = auto()

    # TVTC sequence command #
    ENV_SERVO = auto()
    ENV_OSC = auto()
    ENV_TEMP_START = auto()
    ENV_TEMP_LIMIT = auto()
    WAIT = auto()
    SET_SERVO = auto()
    SET_TARGET = auto()
    MEASURE = auto()
    SEQUENCE_END = auto()


class MeasSDR:
    def __init__(self):
        super().__init__()
        self.measurement_script = None
        self.visa_rx = None

        self.result_path = './data/result/'

        # sequence managing
        self.cmd_running_state = None
        self.next_cmd_start_time = 0
        self.dynamo_start_time = None
        self.cmd_start_time = None
        self.sequence_start_line = None

        # can default value or prev value
        self.prev_can_cmd = dict()
        self.prev_target_motor_dict = None

        # equip env
        self.env_servo = 'RPM'
        self.servo_ready = False
        self.env_osc = 'RMS'
        self.env_temp_start = None
        self.env_temp_limit = None
        self.env_monitoring = False

        # logging
        self.log_thread = None
        self.log_state = False
        self.log_f = 'temp.xlsx'
        self.wb = None
        self.cur_row = 0
        self.log_list = []
        self.servo_speed_log = None
        self.target1_duty_log = None
        self.target2_duty_log = None
        self.alpha_angle_ch1_log = None

        self.is_stop_meas = False
        self.sample_rate = 0.01

        self.init()

    def init(self):
        if not os.path.exists(self.result_path):
            os.makedirs(self.result_path)

        # Visa Thread
        self.visa_rx = VisaRxTread(self.env_osc)
        if 'POWER' in visa.status.keys() or 'OSCILLOSCOPE' in visa.status.keys():
            self.visa_rx.start()

        # logging thread
        self.log_thread = threading.Thread(target=self.logging_func)
        self.log_thread.daemon = True
        self.log_thread.start()

    def load_script(self, testEnv):
        testEnv = testEnv.fillna('0')
        self.measurement_script = testEnv.values.tolist()

    def stop(self):
        self.cmd_running_state = CmdState.DONE
        self.log_state = False
        self.visa_rx.measure_state = False

        if self.env_servo == 'RPM':
            if canBus.status != CAN_ERR:
                self.set_target_motor_on_off(False)
            time.sleep(0.2)

            if telnet.statis is True:
                telnet.set_servo_speed(0)
                telnet.enable_servo(False)
                telnet.servo_ready = False
            time.sleep(0.2)
        else:
            if telnet.statis is True:
                telnet.set_servo_torque(0)
                telnet.enable_servo(False)
                telnet.servo_ready = False
            time.sleep(0.2)

            if canBus.status != CAN_ERR:
                self.set_target_motor_on_off(False)
            time.sleep(0.2)

        # logging file close
        if self.wb is not None:
            self.log_save()
            try:
                self.wb.close()
            except Exception as e:
                print(e)
            self.wb = None

    def start(self):
        self.cmd_running_state = CmdState.READY
        cur_cmd_line = 0
        cmd_dict = None

        while self.cmd_running_state != CmdState.DONE:
            if self.cmd_running_state == CmdState.TEMP_WAIT_MODE:
                cur_cmd_line = self.sequence_start_line
                self.cmd_running_state = CmdState.READY

            if self.cmd_running_state in [CmdState.READY, CmdState.NEXT]:
                if self.cmd_running_state == CmdState.READY:
                    self.sequence_start_line = cur_cmd_line

                cmd_dict = self.get_cmd_line(cur_cmd_line)
                cur_cmd_line += 1

            if cmd_dict is None or self.is_stop_meas is True:
                break
            elif cmd_dict['CMD'] is None or cmd_dict['CMD'] == '':
                continue
            elif cmd_dict['CMD'] == 'SEQUENCE_END':
                self.stop()
                self.cmd_running_state = CmdState.READY
            else:
                self.next_cmd_start_time = float(cmd_dict['Time[sec]'])
                self.run_cmd_process(cmd_dict)

        self.stop()
        print('>> Test Done\n')

    def log_save(self):
        for log_line in self.log_list:
            for col, item in enumerate(log_line):
                sheet = self.wb.active
                sheet.cell(row=self.cur_row, column=col + 1, value=item)
            self.cur_row += 1
        self.wb.save(self.log_f)
        # Save CSV
        df_meas = pd.DataFrame(self.log_list, columns=MEASURE_LIST)
        df_meas.to_csv(self.log_f.replace('.xlsx', '.csv'), index=False)

        self.log_list.clear()

    def get_cmd_line(self, cur_cmd_line):
        if cur_cmd_line >= len(self.measurement_script):
            self.cmd_running_state = CmdState.DONE
            return None
        else:
            cmd_dict = dict(zip(CMD_LIST, self.measurement_script[cur_cmd_line]))
            return cmd_dict

    def run_cmd_process(self, cmd_dict):
        if self.cmd_running_state == CmdState.READY:
            self.ready_process()
            self.log_setting()

        if self.cmd_running_state == CmdState.RUNNING:
            time_diff = datetime.datetime.now() - self.cmd_start_time

            if self.prev_target_motor_dict is not None:
                if canBus.status != CAN_ERR:
                    self.write_can_msg(self.prev_target_motor_dict)

            if time_diff.total_seconds() >= self.next_cmd_start_time:
                self.cmd_running_state = CmdState.NEXT
        else:
            print("Commmand in Progress: {}\n".format(cmd_dict['CMD']))
            if cmd_dict['CMD'] == 'MEASURE_START':
                self.log_state = True
            elif cmd_dict['CMD'] == 'MEASURE_END':
                self.log_state = False
            elif cmd_dict['CMD'] == 'ENV_SERVO':
                if telnet.status is True:
                    self.env_servo = cmd_dict['Env_Val']
                    telnet.set_servo_mode(self.env_servo)
            elif cmd_dict['CMD'] == 'ENV_OSC':
                self.env_osc = cmd_dict['Env_Val']
            elif cmd_dict['CMD'] == 'ENV_TEMP_START':
                self.env_temp_start = None
                self.env_monitoring = False
                if cmd_dict['Env_Val'] is not None:
                    self.env_temp_limit = cmd_dict['Env_Val']
                    if self.env_temp_limit is not None:
                        self.env_monitoring = True
            elif cmd_dict['CMD'] == 'WAIT':
                pass
            elif cmd_dict['CMD'] == 'SET_SERVO':
                if telnet.status is True:
                    if telnet.servo_ready is False:
                        telnet.servo_ready = True
                        telnet.enable_servo(True)
                    if self.env_servo == 'RPM':
                        telnet.set_servo_speed(int(cmd_dict['Servo_Speed']))
                    else:
                        telnet.set_servo_torque(float(cmd_dict['Servo_Speed']))
                self.servo_speed_log = cmd_dict['Servo_Speed']
            elif cmd_dict['CMD'] == 'SET_TARGET':
                if canBus.status != CAN_ERR:
                    self.write_can_msg(cmd_dict)
            elif cmd_dict['CMD'] == 'STOP_SERVO':
                if telnet.status is True:
                    if self.env_servo == 'RPM':
                        telnet.set_servo_speed(0)
                    else:
                        telnet.set_servo_torque(0)
                self.servo_val = 0
            elif cmd_dict['CMD'] == 'SAMPLE_RATE':
                self.sample_rate = float(cmd_dict['Env_Val'])

            if cmd_dict['CMD'] == 'SET_TARGET':
                self.prev_target_motor_dict = cmd_dict
            else:
                self.prev_target_motor_dict = None

            self.cmd_running_state = CmdState.RUNNING
            self.cmd_start_time = datetime.datetime.now()

    def ready_process(self):
        if canBus.status != CAN_ERR:
            self.set_target_motor_on_off(True)

        self.visa_rx.measure_state = True

        self.dynamo_start_time = datetime.datetime.now()

    def set_target_motor_on_off(self, on_off):
        can_msg = dict(zip(CMD_LIST, ['0'] * len(CMD_LIST)))

        can_msg['nMotCtrlAbsLimCh1'] = '5000'
        can_msg['outMotReqCh2'] = '0'
        can_msg['stMotOffReqCh2'] = '0'
        can_msg['stMotCtrlTarCh2'] = '0'
        can_msg['nMotCtrlAbsLimCh2'] = '5000'
        can_msg['Alpha_AngleCh1'] = '0'

        if on_off is True:
            PWM_ON_VAL = '1'
            can_msg['stMotOffReqCh1'] = '170'
            can_msg['stMotCtrlTarCh1'] = '17'
            can_msg['outMotReqCh1'] = PWM_ON_VAL
            self.write_can_msg(cmd_dict=can_msg)
        else:
            can_msg['outMotReqCh1'] = '0'
            can_msg['stMotCtrlTarCh1'] = '0'
            if self.is_stop_meas:
                can_msg['stMotOffReqCh1'] = '85'
            else:
                can_msg['stMotOffReqCh1'] = '170'
            self.write_can_msg(cmd_dict=can_msg)

        time.sleep(0.1)
        self.prev_can_cmd = can_msg

    def log_setting(self):
        print(">> Log Setting\n")
        self.wb = openpyxl.Workbook()
        sheet = self.wb.active
        sheet.cell(row=1, column=1, value='Automation Measure')
        sheet.cell(row=2, column=1, value='TVTC')
        sheet.cell(row=3, column=1, value='1')
        sheet.cell(row=4, column=1, value='Sample Rate(Hz) : ')
        sheet.cell(row=4, column=2, value=1/self.sample_rate)
        sheet.cell(row=5, column=1, value='Result')
        sheet.cell(row=6, column=1, value='AI_Data')

        for col, item in enumerate(MEASURE_LIST):
            sheet.cell(row=7, column=col + 1, value=item)

        now = datetime.datetime.now()
        now = now.strftime("%Y_%m_%d_%H_%M_%S_%f")
        self.log_f = f"{self.result_path}{now}.xlsx"

        self.wb.save(self.log_f)
        self.cur_row = 8
        self.log_list.clear()

    def logging_func(self):
        while True:
            if self.log_state is True:
                current_time = datetime.datetime.now()
                time_diff = current_time - self.dynamo_start_time

                data_time = current_time.strftime("%Y-%m-%d %H:%M:%S")
                running_time = time_diff.total_seconds()

                can_rx_msg = self.read_can_msg()
                can_dynamo_msg = self.read_dynamo_can_msg()

                # from dynamo can
                dynamo_torque = None
                friction_torque = None

                if self.target1_duty_log is not None and self.target1_duty_log != 0:
                    dynamo_torque = can_dynamo_msg['Dynamo_Torque']
                else:
                    friction_torque = can_dynamo_msg['Dynamo_Torque']

                kt = 0
                iMotFdBckCh1 = can_rx_msg['ExtIn_iMotFdBckCh1']
                if iMotFdBckCh1 != 0 and dynamo_torque is not None:
                    if self.target1_duty_log is not None and self.target1_duty_log != 0:
                        kt = dynamo_torque / iMotFdBckCh1

                # Measurement
                measure_val = [data_time, round(running_time, 2), self.servo_speed_log, self.target1_duty_log,
                               self.target2_duty_log, self.alpha_angle_ch1_log, None, kt, self.visa_rx.data_dict['dc_current'],
                               self.visa_rx.data_dict['dc_voltage'], dynamo_torque, friction_torque, can_dynamo_msg['Dynamo_Speed'],
                               can_dynamo_msg['Dynamo_Volt'], can_dynamo_msg['Dynamo_Curr'], self.visa_rx.data_dict['current_p1'],
                               self.visa_rx.data_dict['current_p2'], self.visa_rx.data_dict['current_p3'], self.visa_rx.data_dict['voltage_p1'],
                               self.visa_rx.data_dict['voltage_p2'], self.visa_rx.data_dict['voltage_p3'], self.visa_rx.data_dict['power_ratio_p1'],
                               self.visa_rx.data_dict['power_ratio_p2'], self.visa_rx.data_dict['power_ratio_p3'],
                               can_rx_msg['ExtIn_tmpPCBCh1'], can_rx_msg['ExtIn_iMotFdBckCh1'], can_rx_msg['ExtIn_uBat'],
                               can_rx_msg['ExtIn_stMotOpModCh1'], can_rx_msg['ExtIn_nMotFdBckCh1']]
                self.log_list.append(measure_val)
            time.sleep(self.sample_rate)

    def write_can_msg(self, cmd_dict: dict):
        # DYNO
        can_msg = dict()
        can_msg['DYNO_Bitset01'] = int(cmd_dict['nMotCtrlAbsLimCh1'])
        can_msg['DYNO_Bitset02'] = int(cmd_dict['nMotCtrlAbsLimCh2'])
        can_msg['DYNO_Bitset03'] = int(cmd_dict['Alpha_AngleCh1'])
        can_msg['DYNO_Bitset04'] = int(cmd_dict['CurrTrig1stCh1'])
        can_msg['DYNO_Bitset05'] = int(cmd_dict['CurrTrig2ndCh1'])
        can_msg['DYNO_Bitset06'] = int(cmd_dict['Alpha_AngleCh2'])
        can_msg['DYNO_Bitset07'] = int(cmd_dict['CurrTrig1stCh2'])
        can_msg['DYNO_Bitset08'] = int(cmd_dict['CurrTrig2ndCh2'])
        canBus.msg_write_by_frame(frame_name='DYNO_10ms', frame_msg=can_msg, time_out=0.1)

    def read_can_msg(self):
        can_rx_msg = dict()
        if canBus.status != CAN_ERR:
            can_0x640 = canBus.msg_read_id(can_id=0x640)
            can_0x644 = canBus.msg_read_id(can_id=0x644)
            can_0x645 = canBus.msg_read_id(can_id=0x645)
            can_0x646 = canBus.msg_read_id(can_id=0x646)
            can_0x647 = canBus.msg_read_id(can_id=0x647)
        else:
            can_0x640 = None
            can_0x644 = None
            can_0x645 = None
            can_0x646 = None
            can_0x647 = None

        if can_0x640 is None:
            can_rx_msg['ExtIn_tmpPCBch1'] = 0
            can_rx_msg['ExtIn_tmpPCBch2'] = 0
        else:
            can_rx_msg['ExtIn_tmpPCBch1'] = can_0x640['ExtIn_tmpPCBch1']
            can_rx_msg['ExtIn_tmpPCBch2'] = can_0x640['ExtIn_tmpPCBch2']

        return can_rx_msg

    def read_dynamo_can_msg(self):
        can_rx_msg = dict()
        if canBus.status != CAN_ERR:
            can_0x91 = canBus.msg_read_id(can_id=0x91)
        else:
            can_0x91 = None

        return can_rx_msg

class VisaRxTread(Thread):
    def __init__(self, env_osc):
        super().__init__()
        self.measure_state = False
        self.env_osc = env_osc
        self.data_dict = {
            'current_p1': 0.0,
            'current_p2': 0.0,
            'current_p3': 0.0,
            'voltage_p1': None,
            'voltage_p2': None,
            'voltage_p3': None,
            'power_ratio_p1': None,
            'power_ratio_p2': None,
            'power_ratio_p3': None,
            'dc_voltage': 0.0,
            'dc_current': 0.0
        }

        self.power_status = False
        self.osc_status = False
        self.update_visa_status()

    def update_visa_status(self):
        self.power_status = False
        if 'POWER' in visa.status.keys():
            if visa.status['POWER'] is True:
                self.power_status = True

        self.osc_status = False
        if 'OSCILLOSCOPE' in visa.status.keys():
            if visa.status['OSCILLOSCOPE'] is True:
                self.osc_status = True

    def run(self):
        while True:
            if self.measure_state is True:
                self.read_oscilloscope()
                self.read_power()
            time.sleep(0.00001)

    def read_oscilloscope(self):
        if self.osc_status is True:
            current_p1 = 'Error - timeout'
            current_p2 = 'Error - timeout'
            current_p3 = 'Error - timeout'

            if self.env_osc == 'MAX':
                current_p1 = visa.read('OSCILLOSCOPE', "C1:PAVA? MAX")
                current_p2 = visa.read('OSCILLOSCOPE', "C2:PAVA? MAX")
                current_p3 = visa.read('OSCILLOSCOPE', "C3:PAVA? MAX")
            elif self.env_osc == 'RMS':
                current_p1 = visa.read('OSCILLOSCOPE', "C1:PAVA? RMS")
                current_p2 = visa.read('OSCILLOSCOPE', "C2:PAVA? RMS")
                current_p3 = visa.read('OSCILLOSCOPE', "C3:PAVA? RMS")

            if current_p1 != 'Error - timeout':
                current_p1_val = current_p1.split(',')[-2].split()[0]
                current_p2_val = current_p2.split(',')[-2].split()[0]
                current_p3_val = current_p3.split(',')[-2].split()[0]

                if current_p3_val != 'UNDEF':
                    self.data_dict['current_p1'] = float(current_p1_val)
                    self.data_dict['current_p2'] = float(current_p2_val)
                    self.data_dict['current_p3'] = float(current_p3_val)

    def read_power(self):
        if self.power_status is True:
            str_volt, str_curr = visa.meas_volt_curr('POWER')
            self.data_dict['dc_voltage'] = float(str_volt)
            self.data_dict['dc_current'] = float(str_curr)
